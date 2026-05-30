"""导出模块：支持 Excel 和 JSON 格式"""
import json
import io
import logging
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from app.models import FileResult, GlobalStats

logger = logging.getLogger(__name__)


def export_to_excel(results: List[FileResult], stats: GlobalStats) -> bytes:
    """生成Excel文件内容（内存），返回bytes"""
    logger.info("开始生成 Excel，结果数：%d", len(results))
    wb = Workbook()

    # === Sheet 1: 文件识别明细 ===
    ws1 = wb.active
    ws1.title = "识别明细"
    headers = ["文件名", "文档类型", "险种类别", "险种大类", "保险公司", "保单号", "异常标记", "状态", "涉敏人数", "涉敏人员"]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        ws1.cell(row=row_idx, column=1, value=r.filename)
        ws1.cell(row=row_idx, column=2, value=r.document_type_display)
        ws1.cell(row=row_idx, column=3, value=r.insurance_category_display)
        ws1.cell(row=row_idx, column=4, value=r.insurance_branch_display)
        ws1.cell(row=row_idx, column=5, value=r.insurance_company)
        ws1.cell(row=row_idx, column=6, value=r.policy_number)
        ws1.cell(row=row_idx, column=7, value=r.anomaly)
        ws1.cell(row=row_idx, column=8, value=r.status)
        ws1.cell(row=row_idx, column=9, value=r.sensitive_count)

        # 涉敏人员列：展示姓名和角色
        persons_text = ""
        if r.persons:
            persons_text = "; ".join([
                f"{p.name or '(匿名)'}({p.role_display})"
                for p in r.persons
            ])
        ws1.cell(row=row_idx, column=10, value=persons_text)

    # === Sheet 2: 统计 ===
    ws2 = wb.create_sheet("敏感信息统计")
    row = 1
    ws2.cell(row=row, column=1, value="统计项目").font = header_font
    ws2.cell(row=row, column=2, value="数值").font = header_font
    row += 1
    ws2.cell(row=row, column=1, value="总文件数")
    ws2.cell(row=row, column=2, value=stats.total_files)
    row += 1
    ws2.cell(row=row, column=1, value="涉敏文件数")
    ws2.cell(row=row, column=2, value=stats.sensitive_files)
    row += 1
    ws2.cell(row=row, column=1, value="涉敏人数（去重）")
    ws2.cell(row=row, column=2, value=stats.global_unique_persons)
    row += 1
    ws2.cell(row=row, column=1, value="非涉敏文件数")
    ws2.cell(row=row, column=2, value=stats.non_sensitive_files)
    row += 1
    ws2.cell(row=row, column=1, value="人身险涉敏文件数")
    ws2.cell(row=row, column=2, value=stats.life_sensitive_files)
    row += 1
    ws2.cell(row=row, column=1, value="人身险涉敏人数（去重）")
    ws2.cell(row=row, column=2, value=stats.life_unique_persons)
    row += 1
    ws2.cell(row=row, column=1, value="财产险文件数")
    ws2.cell(row=row, column=2, value=stats.property_files)
    row += 1
    ws2.cell(row=row, column=1, value="财产险涉敏人数")
    ws2.cell(row=row, column=2, value=stats.property_sensitive_persons)
    row += 1
    ws2.cell(row=row, column=1, value="异常文件数")
    ws2.cell(row=row, column=2, value=stats.anomaly_files)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    data = buf.getvalue()
    logger.info("Excel 生成完成，结果数：%d，字节数：%d", len(results), len(data))
    return data


def export_to_json(results: List[FileResult], stats: GlobalStats) -> str:
    """生成JSON字符串（用于下载）"""
    logger.info("开始生成 JSON，结果数：%d", len(results))
    data = {
        "global_stats": {
            "total_files": stats.total_files,
            "sensitive_files": stats.sensitive_files,
            "global_unique_persons": stats.global_unique_persons,
            "non_sensitive_files": stats.non_sensitive_files,
            "life_sensitive_files": stats.life_sensitive_files,
            "life_unique_persons": stats.life_unique_persons,
            "property_files": stats.property_files,
            "property_sensitive_persons": stats.property_sensitive_persons,
            "anomaly_files": stats.anomaly_files,
        },
        "details": [
            {
                "filename": r.filename,
                "is_insurance_related": r.is_insurance_related,
                "document_type": r.document_type,
                "document_type_display": r.document_type_display,
                "insurance_category": r.insurance_category,
                "insurance_category_display": r.insurance_category_display,
                "insurance_branch": r.insurance_branch,
                "insurance_branch_display": r.insurance_branch_display,
                "anomaly": r.anomaly,
                "insurance_company": r.insurance_company,
                "policy_number": r.policy_number,
                "status": r.status,
                "sensitive_count": r.sensitive_count,
                "persons": [
                    {
                        "name": p.name,
                        "role": p.role,
                        "role_display": p.role_display,
                        "details": [
                            {
                                "type": d.type,
                                "value": d.value,
                                "raw_label": d.raw_label,
                            }
                            for d in p.details
                        ],
                    }
                    for p in r.persons
                ],
                "error_message": r.error_message,
            }
            for r in results
        ],
    }
    json_str = json.dumps(data, ensure_ascii=False, indent=2)
    logger.info("JSON 生成完成，结果数：%d，字符数：%d", len(results), len(json_str))
    return json_str
