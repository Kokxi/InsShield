"""测试导出模块"""
import io
import json
import pytest
from app.exporter import export_to_excel, export_to_json
from app.models import FileResult, GlobalStats, PersonModel


def make_result(name: str, category: str = "life") -> FileResult:
    return FileResult(
        filename=f"{name}.pdf",
        is_insurance_related=True,
        insurance_category=category,
        insurance_category_display={"life": "人寿险", "health": "健康险", "accident": "意外险", "car": "车险", "property": "财产险"}.get(category, "未知"),
        insurance_branch={"life": "life", "health": "life", "accident": "life", "car": "property", "property": "property"}.get(category, "unknown"),
        insurance_branch_display={"life": "人身保险", "property": "财产保险"}.get(category, "未知"),
        persons=[PersonModel(name=name, role="insured", role_display="被保人")],
        sensitive_count=1,
        status="ok",
    )


class TestExportToJson:
    def test_export_structure(self):
        results = [make_result("李四"), make_result("张三")]
        stats = GlobalStats(life_sensitive_files=2, life_unique_persons=2)
        output = export_to_json(results, stats)
        data = json.loads(output)
        assert data["global_stats"]["life_sensitive_files"] == 2
        assert len(data["details"]) == 2
        assert data["details"][0]["filename"] == "李四.pdf"
        assert data["details"][1]["filename"] == "张三.pdf"

    def test_empty_export(self):
        output = export_to_json([], GlobalStats())
        data = json.loads(output)
        assert data["global_stats"]["total_files"] == 0
        assert data["details"] == []

    def test_json_is_readable(self):
        results = [make_result("李四")]
        stats = GlobalStats(life_sensitive_files=1, life_unique_persons=1)
        output = export_to_json(results, stats)
        assert "李四" in output

    def test_insurance_category_in_detail(self):
        """detail中应包含insurance_category字段"""
        results = [make_result("李四")]
        stats = GlobalStats(life_sensitive_files=1, life_unique_persons=1)
        output = export_to_json(results, stats)
        data = json.loads(output)
        assert data["details"][0]["insurance_category"] == "life"

    def test_insurance_branch_in_detail(self):
        """detail中应包含insurance_branch字段"""
        results = [make_result("李四")]
        stats = GlobalStats()
        output = export_to_json(results, stats)
        data = json.loads(output)
        assert data["details"][0]["insurance_branch"] == "life"
        assert data["details"][0]["insurance_branch_display"] == "人身保险"

    def test_anomaly_in_detail(self):
        """detail中应包含anomaly字段"""
        results = [FileResult(
            filename="异常.pdf", is_insurance_related=True,
            insurance_category="car", insurance_category_display="车险",
            insurance_branch="property", insurance_branch_display="财产保险",
            anomaly="财产险多人",
            persons=[PersonModel(name="张三"), PersonModel(name="李四")],
            sensitive_count=2, status="ok",
        )]
        stats = GlobalStats(anomaly_files=1)
        output = export_to_json(results, stats)
        data = json.loads(output)
        assert data["details"][0]["anomaly"] == "财产险多人"
        assert data["global_stats"]["anomaly_files"] == 1


class TestExportToExcel:
    def test_export_creates_workbook(self):
        results = [make_result("李四")]
        stats = GlobalStats(life_sensitive_files=1, life_unique_persons=1)
        data = export_to_excel(results, stats)
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_export_multiple_rows(self):
        results = [make_result("李四"), make_result("张三")]
        stats = GlobalStats(life_sensitive_files=2, life_unique_persons=2)
        data = export_to_excel(results, stats)
        assert isinstance(data, bytes)
        assert len(data) > 100

    def test_export_empty(self):
        data = export_to_excel([], GlobalStats())
        assert isinstance(data, bytes)

    def test_has_two_sheets(self):
        from openpyxl import load_workbook
        results = [make_result("李四")]
        stats = GlobalStats(life_sensitive_files=1, life_unique_persons=1)
        data = export_to_excel(results, stats)
        wb = load_workbook(io.BytesIO(data))
        assert "识别明细" in wb.sheetnames
        assert "敏感信息统计" in wb.sheetnames

    def test_insurance_branch_column(self):
        """检查险种大类列存在且有值"""
        from openpyxl import load_workbook
        results = [make_result("李四")]
        stats = GlobalStats(life_sensitive_files=1, life_unique_persons=1)
        data = export_to_excel(results, stats)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["识别明细"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert "险种大类" in headers
        branch_col = headers.index("险种大类") + 1
        assert ws.cell(row=2, column=branch_col).value == "人身保险"

    def test_anomaly_column(self):
        """检查异常标记列存在"""
        from openpyxl import load_workbook
        results = [make_result("李四")]
        stats = GlobalStats()
        data = export_to_excel(results, stats)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["识别明细"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert "异常标记" in headers

    def test_stats_sheet_has_branch_rows(self):
        """统计Sheet包含分支维度行"""
        from openpyxl import load_workbook
        results = [make_result("李四")]
        stats = GlobalStats(
            life_sensitive_files=1, life_unique_persons=1,
            property_files=0, property_sensitive_persons=0,
            anomaly_files=0,
        )
        data = export_to_excel(results, stats)
        wb = load_workbook(io.BytesIO(data))
        ws = wb["敏感信息统计"]
        rows = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value for r in range(1, 20)}
        assert "人身险涉敏文件数" in rows
        assert "财产险文件数" in rows
        assert "异常文件数" in rows
